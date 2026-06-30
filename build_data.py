"""Export shipping price data from Excel to JSON. Spouštějte lokálně před push na GitHub."""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

APP_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = Path(r"e:\AI\vnitro dopravy\Kostelec\Vnitro_Kostelec.xlsx")
DEFAULT_OUT = APP_DIR / "data" / "shipping.json"


def parse_zony(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["ZONY"]
    entries: list[dict] = []
    for r in range(3, ws.max_row + 1):
        psc_od = ws.cell(r, 1).value
        psc_do = ws.cell(r, 2).value
        zone = ws.cell(r, 3).value
        city = ws.cell(r, 4).value
        if psc_od is None or psc_do is None or zone is None:
            continue
        try:
            entries.append(
                {
                    "psc_od": int(psc_od),
                    "psc_do": int(psc_do),
                    "zone": int(zone),
                    "city": str(city) if city else "",
                }
            )
        except (TypeError, ValueError):
            continue
    return entries


def parse_raben(wb: openpyxl.Workbook) -> dict:
    ws = wb["Raben_Kostelec"]

    fuel = ws.cell(20, 11).value
    customer = ws.cell(21, 11).value
    multiplier = (1 + (float(fuel) if isinstance(fuel, (int, float)) else 0)) * (
        1 + (float(customer) if isinstance(customer, (int, float)) else 0)
    )

    weights: list[int] = []
    prices: dict[str, dict[str, float | None]] = {"1": {}, "2": {}, "3": {}}
    zone_cols = {1: 4, 2: 5, 3: 6}
    weight_row_start = 14
    base_row_start = 59

    for r in range(weight_row_start, ws.max_row + 1):
        w = ws.cell(r, 2).value
        if isinstance(w, str):
            break
        if not isinstance(w, (int, float)):
            continue
        weight = int(w)
        if weight in weights:
            continue
        base_row = base_row_start + (r - weight_row_start)
        row_prices: dict[str, float | None] = {}
        for zone_num, col in zone_cols.items():
            val = ws.cell(r, col).value
            if not isinstance(val, (int, float)):
                base_val = ws.cell(base_row, col).value
                if isinstance(base_val, (int, float)):
                    val = float(base_val) * multiplier
            row_prices[str(zone_num)] = float(val) if isinstance(val, (int, float)) else None
        if any(v is not None for v in row_prices.values()):
            weights.append(weight)
            for zone_key, price in row_prices.items():
                prices[zone_key][str(weight)] = price

    weights.sort()

    psc_ranges = []
    for r in range(26, ws.max_row + 1):
        psc_od = ws.cell(r, 9).value
        psc_do = ws.cell(r, 10).value
        zone = ws.cell(r, 11).value
        if isinstance(psc_od, str) or isinstance(psc_do, str):
            continue
        if not isinstance(psc_od, (int, float)) or not isinstance(psc_do, (int, float)):
            continue
        if not isinstance(zone, (int, float)):
            continue
        psc_ranges.append(
            {"psc_od": int(psc_od), "psc_do": int(psc_do), "zone": int(zone)}
        )

    return {"psc_ranges": psc_ranges, "weights": weights, "prices": prices}


def parse_dnp(wb: openpyxl.Workbook) -> dict:
    ws = wb["DNP_Kostelec"]

    weights: list[int] = []
    for c in range(2, ws.max_column + 1):
        w = ws.cell(2, c).value
        if isinstance(w, (int, float)):
            weights.append(int(w))

    prices: dict[str, dict[str, float | None]] = {}
    for r in range(3, ws.max_row + 1):
        zone = ws.cell(r, 1).value
        if not isinstance(zone, (int, float)) or zone > 100:
            break
        zone_key = str(int(zone))
        prices[zone_key] = {}
        for c, w in enumerate(weights, start=2):
            val = ws.cell(r, c).value
            prices[zone_key][str(w)] = float(val) if isinstance(val, (int, float)) else None

    psc_ranges: list[dict] = []
    for r in range(18, ws.max_row + 1):
        psc_od = ws.cell(r, 1).value
        if isinstance(psc_od, str) and "počítat" in psc_od.lower():
            break
        psc_do = ws.cell(r, 2).value
        zone = ws.cell(r, 5).value
        if not isinstance(psc_od, (int, float)) or not isinstance(psc_do, (int, float)):
            continue
        if not isinstance(zone, (int, float)):
            continue
        psc_ranges.append(
            {
                "psc_od": int(psc_od),
                "psc_do": int(psc_do),
                "zone": int(zone),
                "okres": str(ws.cell(r, 4).value or ""),
            }
        )

    return {"psc_ranges": psc_ranges, "weights": weights, "prices": prices}


def parse_vnitro(wb: openpyxl.Workbook) -> dict:
    ws = wb["Vnitro_Kostelec"]

    weight_limits = [2000, 3000, 6000, 8000, 12000, 999999]
    limit_cols = [7, 12, 17, 22, 27, 32]
    minimum_cols = [6, 11, 16, 21, 26, 31]

    prices: dict[str, dict] = {}
    for r in range(6, ws.max_row + 1):
        pasmo = ws.cell(r, 1).value
        if not isinstance(pasmo, (int, float)):
            continue
        pasmo_key = str(int(pasmo))
        km_range = ws.cell(r, 2).value
        prices[pasmo_key] = {"km_range": str(km_range) if km_range else ""}
        for limit, limit_col, min_col in zip(weight_limits, limit_cols, minimum_cols):
            limit_key = "999999" if limit == 999999 else str(limit)
            limit_val = ws.cell(r, limit_col).value
            min_val = ws.cell(r, min_col).value
            prices[pasmo_key][limit_key] = {
                "limit": float(limit_val) if isinstance(limit_val, (int, float)) else None,
                "minimum": float(min_val) if isinstance(min_val, (int, float)) else None,
            }

    return {"weight_limits": weight_limits[:-1] + [999999], "prices": prices}


def build_data(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    return {
        "meta": {
            "source": xlsx_path.name,
            "carriers": ["Raben", "DNP", "Vnitro"],
        },
        "zony": parse_zony(wb),
        "raben": parse_raben(wb),
        "dnp": parse_dnp(wb),
        "vnitro": parse_vnitro(wb),
    }


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT

    if not xlsx_path.exists():
        print(f"Chyba: soubor {xlsx_path} neexistuje.", file=sys.stderr)
        sys.exit(1)

    data = build_data(xlsx_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Export dokoncen: {out_path}")
    print(f"  ZONY: {len(data['zony'])} intervalu PSC")
    print(f"  Raben: {len(data['raben']['weights'])} hmotnostnich pasem")
    print(f"  DNP: {len(data['dnp']['prices'])} zon")
    print(f"  Vnitro: {len(data['vnitro']['prices'])} pasem")


if __name__ == "__main__":
    main()
